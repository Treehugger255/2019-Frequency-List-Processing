import spacy
import openpyxl
from collections import Counter
import math

writeBook = input("What file to write to? ")
openFile = input("What file do you want to open? ")

workbookWrite = openpyxl.Workbook()
sheet = workbookWrite.active
sheet_title = "Sheet1"

def writeData(frqDict, docLength):
    lemmasPOSList = list(frqDict.keys())
    separateCounter = 0
    while separateCounter < len(lemmasPOSList):
        currentRow = separateCounter + 2
        currentColumn = 1
        lemmaPOS = lemmasPOSList[separateCounter]
        lemmaList = lemmaPOS.split("_")
        frq = frqDict[lemmaPOS]
        fpmw = float(frqDict[lemmaPOS]) * (10 ** 6) / float(docLength)
        zipf = math.log(fpmw * 1000, 10)
        for j in [lemmaList[0], lemmaList[1], frq, round((fpmw), 2), round((zipf), 6)]:
            sheet.cell(row=currentRow, column= currentColumn).value = j
            currentColumn += 1
        separateCounter += 1

nlp = spacy.load("en")

f = open(openFile, "r")
file = f.read()
doc = nlp(file)

lemmaList = []
for word in doc:
    if word.pos_ != "PUNCT" and word.lemma_ != "\n" and word.pos_ != "SPACE":
        lemmaList.extend([word.lemma_ + "_" + word.pos_])

headerCounter = 1

for header in ["Word", "PoS", "FREQcount", "SUBTLWF", "Zipf-value"]: # Creates the header
    headerCell = sheet.cell(row = 1, column = headerCounter)
    headerCell.value = header
    headerCounter += 1

#lemmaList = lemmaDict.keys()
#print(lemmaList)
totalTokens = len(lemmaList)
counts = Counter(lemmaList)
#print(percentTotal(counts, totalTokens, list2))
writeData(counts, totalTokens)
#print(counts["be_VERB"])
#print(counts)

workbookWrite.save(writeBook)
f.close()